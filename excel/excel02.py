from openpyxl import load_workbook

filesheet="./demosheet.xlsx"

#creamos el objeto load_workbook
wb = load_workbook(filesheet)

#seleccionamos el archivo
sheet = wb.active

#ingresamos el valor 56 en la celda 'A1'
sheet['A1']=56

#ingresamos el valor 1845 en la celda 'B3'
sheet['B3']=1845

#guardamos el archivo con los cambios
wb.save(filesheet)
