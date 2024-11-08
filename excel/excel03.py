from openpyxl import load_workbook
import os

filesheet="./demosheet.xlsx"

#creamos el objeto load_workbook
wb = load_workbook(filesheet)

#seleccionamos el archivo
sheet = wb.active

# escribimos los datos en sus respectivas celdas
datos = [('id','nombre','edad'),
         (0,'alex',18),
         (2,'gustaf',16),
         (4, 'tre',21)]

for row in datos:
    sheet.append(row)

wb.save(filesheet)

